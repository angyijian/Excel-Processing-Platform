"""
Compare a Reorder workbook against a Purchase workbook.

Matching logic:
  1. Exact match on SKUCODE + ETD  -> Unchanged / Qty Adjusted
  2. Leftover rows grouped by SKUCODE only:
       - exactly 1 leftover in Reorder <-> exactly 1 leftover in Purchase
             -> ETD Changed  (also flags qty change if it moved too)
       - SKUCODE has leftovers on both sides but not 1-to-1
             -> Needs Review (ambiguous, don't guess)
       - leftover only in Reorder -> Cancelled
       - leftover only in Purchase -> Added

Usage:
    python compare_po.py reorder.xlsx purchase.xlsx
"""
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_NAME = "Compare PO Files"
REQUIRED_FILE_COUNT = 2

SKU = "SKUCODE"
QTY = "总数量\n(TOTAL QTY)"
SUPPLIER = "SUPPLIER"
ETD = "ETD\n(送货日期)"
ETA = "ETA\n(到厂日期)"
COST = "Order Cost"

STATUS_COLORS = {
    "Cancelled":      "FFC7CE",  # red
    "Added":          "C6EFCE",  # green
    "Qty Adjusted":   "FFEB9C",  # yellow
    "ETD Changed":    "9BC2E6",  # blue
    "ETD + Qty Changed": "B4A7D6",  # purple
    "Needs Review":   "F4B084",  # orange
    "Unchanged":      "FFFFFF",  # white
}


def validate_input_files(file_paths):
    if not file_paths or len(file_paths) != 2:
        raise ValueError("This script requires exactly 2 Excel files: Reorder and Purchase.")

    reorder_path = None
    purchase_path = None

    for path in file_paths:
        if not os.path.exists(path):
            raise FileNotFoundError(f"File not found: {path}")

        filename = os.path.basename(path).lower()
        if "reorder" in filename:
            reorder_path = path
        elif "purchase" in filename:
            purchase_path = path

    if reorder_path is None or purchase_path is None:
        raise ValueError("Expected file names to include 'Reorder' and 'Purchase'.")

    return reorder_path, purchase_path


def load(reorder_path, purchase_path):
    reorder_df = pd.read_excel(reorder_path, sheet_name=0)
    purchase_df = pd.read_excel(purchase_path, sheet_name=0)
    return reorder_df, purchase_df


def compare(r: pd.DataFrame, p: pd.DataFrame) -> pd.DataFrame:
    r = r.copy()
    p = p.copy()
    r["_ridx"] = r.index
    p["_pidx"] = p.index

    r_key = list(zip(r[SKU], r[ETD]))
    p_key = list(zip(p[SKU], p[ETD]))
    r["_key"] = r_key
    p["_key"] = p_key

    p_lookup = {k: i for i, k in zip(p["_pidx"], p["_key"])}

    rows = []
    matched_r = set()
    matched_p = set()

    # Step 1: exact SKUCODE + ETD match
    for _, rrow in r.iterrows():
        k = rrow["_key"]
        if k in p_lookup:
            pidx = p_lookup[k]
            prow = p.loc[p["_pidx"] == pidx].iloc[0]
            matched_r.add(rrow["_ridx"])
            matched_p.add(pidx)
            status = "Unchanged" if rrow[QTY] == prow[QTY] else "Qty Adjusted"
            rows.append(build_row(status, rrow, prow))

    # Step 2: leftovers, group by SKUCODE only
    r_left = r[~r["_ridx"].isin(matched_r)]
    p_left = p[~p["_pidx"].isin(matched_p)]

    r_groups = r_left.groupby(SKU)
    p_groups = dict(list(p_left.groupby(SKU)))

    handled_p_groups = set()

    for sku, rgrp in r_groups:
        pgrp = p_groups.get(sku)
        if pgrp is None or len(pgrp) == 0:
            # only in Reorder -> Cancelled (all leftover rows for this SKU)
            for _, rrow in rgrp.iterrows():
                rows.append(build_row("Cancelled", rrow, None))
        elif len(rgrp) == 1 and len(pgrp) == 1:
            rrow = rgrp.iloc[0]
            prow = pgrp.iloc[0]
            qty_same = rrow[QTY] == prow[QTY]
            status = "ETD Changed" if qty_same else "ETD + Qty Changed"
            rows.append(build_row(status, rrow, prow))
            handled_p_groups.add(sku)
        else:
            # Multiple leftover rows for this SKU on both sides.
            # Pair rows that share the exact same quantity (sorted by ETD) as
            # "ETD Changed" -- these are just re-dated duplicates, not ambiguous.
            # Only rows whose quantity has no remaining match stay "Needs Review".
            r_by_qty = {}
            for _, rrow in rgrp.sort_values(ETD).iterrows():
                r_by_qty.setdefault(rrow[QTY], []).append(rrow)
            p_by_qty = {}
            for _, prow in pgrp.sort_values(ETD).iterrows():
                p_by_qty.setdefault(prow[QTY], []).append(prow)

            for qty_val in list(r_by_qty.keys()):
                if qty_val in p_by_qty:
                    rlist = r_by_qty[qty_val]
                    plist = p_by_qty[qty_val]
                    n = min(len(rlist), len(plist))
                    for i in range(n):
                        rows.append(build_row("ETD Changed", rlist[i], plist[i]))
                    r_by_qty[qty_val] = rlist[n:]
                    p_by_qty[qty_val] = plist[n:]

            for qty_val, rlist in r_by_qty.items():
                for rrow in rlist:
                    rows.append(build_row("Needs Review", rrow, None, note="Multiple leftover rows for this SKU, quantity has no exact match on Purchase side"))
            for qty_val, plist in p_by_qty.items():
                for prow in plist:
                    rows.append(build_row("Needs Review", None, prow, note="Multiple leftover rows for this SKU, quantity has no exact match on Reorder side"))
            handled_p_groups.add(sku)

    # SKUs only in Purchase leftovers (Added) -- not already handled above
    for sku, pgrp in p_groups.items():
        if sku in handled_p_groups:
            continue
        if sku not in r_groups.groups:
            for _, prow in pgrp.iterrows():
                rows.append(build_row("Added", None, prow))

    result = pd.DataFrame(rows)
    order = {"Cancelled": 0, "Added": 1, "Needs Review": 2, "ETD + Qty Changed": 3,
             "ETD Changed": 4, "Qty Adjusted": 5, "Unchanged": 6}
    result["_sort"] = result["Status"].map(order)
    result = result.sort_values(["_sort", "SKUCODE"]).drop(columns="_sort").reset_index(drop=True)
    return result


def build_row(status, rrow, prow, note=""):
    def g(row, col):
        return row[col] if row is not None else None

    def total_amount(row):
        if row is None:
            return None
        cost = row[COST]
        qty = row[QTY]
        if pd.isna(cost) or pd.isna(qty):
            return None
        return cost * qty

    reorder_total = total_amount(rrow)
    purchase_total = total_amount(prow)

    return {
        "Status": status,
        "SKUCODE": g(rrow, SKU) if rrow is not None else g(prow, SKU),
        "SUPPLIER": g(rrow, SUPPLIER) if rrow is not None else g(prow, SUPPLIER),
        "Reorder ETD": g(rrow, ETD),
        "Purchase ETD": g(prow, ETD),
        "Reorder Qty": g(rrow, QTY),
        "Purchase Qty": g(prow, QTY),
        "Qty Diff": (g(prow, QTY) - g(rrow, QTY)) if (rrow is not None and prow is not None) else None,
        "Reorder Cost": g(rrow, COST),
        "Purchase Cost": g(prow, COST),
        "Reorder Total Amount": reorder_total,
        "Purchase Total Amount": purchase_total,
        "Amount Diff": (purchase_total - reorder_total)
        if reorder_total is not None and purchase_total is not None else None,
        "Note": note,
    }


def write_output(result: pd.DataFrame, out_path: str):
    cols = ["Status", "SKUCODE", "SUPPLIER", "Reorder ETD", "Purchase ETD",
            "Reorder Qty", "Purchase Qty", "Qty Diff", "Reorder Cost", "Purchase Cost",
            "Reorder Total Amount", "Purchase Total Amount", "Amount Diff", "Note"]
    result = result[cols]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison Summary"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="404040")
    body_font = Font(name="Arial", size=10)

    for c, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_i, row in enumerate(result.itertuples(index=False), start=2):
        status = row.Status
        fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
        for c_i, val in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.font = body_font
            cell.fill = fill
            if cols[c_i - 1] in ("Reorder ETD", "Purchase ETD"):
                if val is not None:
                    cell.number_format = "yyyy-mm-dd"
            elif cols[c_i - 1] in (
                "Reorder Cost", "Purchase Cost", "Reorder Total Amount",
                "Purchase Total Amount", "Amount Diff"
            ):
                if val is not None:
                    cell.number_format = "#,##0.00"

    widths = [16, 12, 12, 14, 14, 12, 12, 10, 14, 14, 20, 20, 14, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(result)+1}"

    # Summary counts sheet
    ws2 = wb.create_sheet("Status Counts")
    counts = result["Status"].value_counts()
    ws2.cell(row=1, column=1, value="Status").font = header_font
    ws2.cell(row=1, column=2, value="Count").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2).fill = header_fill
    for i, (status, cnt) in enumerate(counts.items(), start=2):
        ws2.cell(row=i, column=1, value=status).font = body_font
        ws2.cell(row=i, column=2, value=int(cnt)).font = body_font
        fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
        ws2.cell(row=i, column=1).fill = fill
        ws2.cell(row=i, column=2).fill = fill
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10

    wb.save(out_path)


def process_data(file_paths, update_progress_callback=None):
    reorder_path, purchase_path = validate_input_files(file_paths)

    if update_progress_callback:
        update_progress_callback(0.1, "Loading workbook data...")

    reorder_df, purchase_df = load(reorder_path, purchase_path)

    if update_progress_callback:
        update_progress_callback(0.5, "Comparing records...")

    result = compare(reorder_df, purchase_df)

    out_path = os.path.join(os.getcwd(), "PO_Comparison_Result.xlsx")
    write_output(result, out_path)

    if update_progress_callback:
        update_progress_callback(1.0, "Finished comparison")

    return out_path


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python compare_po.py Reorder.xlsx Purchase.xlsx")

    reorder_path, purchase_path = validate_input_files([sys.argv[1], sys.argv[2]])
    r, p = load(reorder_path, purchase_path)
    result = compare(r, p)
    write_output(result, os.path.abspath(sys.argv[2].replace(os.path.basename(sys.argv[2]), "PO_Comparison_Result.xlsx")))
    print(result["Status"].value_counts())