import os
import re
import datetime
from copy import copy
import openpyxl

# 1. UI Display Name
SCRIPT_NAME = "Merge PO Excel Reports"
REQUIRED_FILE_COUNT = 1
ALLOW_MULTIPLE_FILES = True

# Updated 17 Expected Headers Example List
RAW_EXAMPLE_HEADERS = [
    "SKUCODE",
    "总数量 (TOTAL QTY)",
    "数量 (QTY/BOX)",
    "单件体积 (CBM/CARTON)",
    "件数 (TOTAL CARTON)",
    "总体积 (TOTAL CBM)",
    "Order Cost",
    "W1 Cost",
    "单号 (ORDER NO)",
    "SUPPLIER",
    "订货日期 (ORDER DATE)",
    "ETD (送货日期)",
    "ETA (到厂日期)",
    "Currency",
    "PO Prefix",
    "Header Remark",
    "Special Instruction"
]


def normalize_header(value) -> str:
    """
    Normalizes a header cell value according to standard specification:
    - Converts to upper case string
    - Removes Chinese characters (\u4e00-\u9fff)
    - Strips parentheses
    - Collapses line breaks, tabs, and multiple whitespaces
    - Trims leading/trailing whitespace
    """
    if value is None:
        return ""
    val_str = str(value).upper()
    # Remove Chinese characters
    val_str = re.sub(r'[\u4e00-\u9fff]', '', val_str)
    # Remove parentheses
    val_str = re.sub(r'[\(\)]', ' ', val_str)
    # Replace line breaks, tabs, and multiple whitespace with a single space
    val_str = re.sub(r'[\r\n\t]+', ' ', val_str)
    val_str = re.sub(r'\s+', ' ', val_str)
    return val_str.strip()


EXPECTED_17_HEADERS = [normalize_header(h) for h in RAW_EXAMPLE_HEADERS]
EXPECTED_13_HEADERS = EXPECTED_17_HEADERS[:13]

DATE_HEADER_NAMES = {"ORDER DATE", "ETD", "ETA"}


def copy_cell_style(source_cell, target_cell):
    """Copy visual formatting from one cell to another."""
    if source_cell is None or target_cell is None:
        return

    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def validate_input_files(file_paths: list[str]) -> list[str]:
    """
    Validates input file paths:
    - Ignores temporary Excel files starting with ~$
    - Filters accepted Excel extensions (.xlsx, .xlsm, .xls, .xlsb)
    - Raises ValueError if no valid files remain
    """
    if not file_paths:
        raise ValueError("No input files provided.")

    valid_paths = []
    valid_extensions = ('.xlsx', '.xlsm', '.xls', '.xlsb')

    for path in file_paths:
        filename = os.path.basename(path)
        if filename.startswith("~$"):
            continue
        if filename.lower().endswith(valid_extensions):
            valid_paths.append(path)

    if not valid_paths:
        raise ValueError("No valid Excel files (.xlsx, .xlsm, .xls, .xlsb) found in the selection.")

    return valid_paths


def find_header_position(sheet) -> tuple[int, int, list[str]] | tuple[None, None, None]:
    """
    Scans Row 1 and Row 2 to detect starting position of matching PO header structure.
    Returns (header_row, start_column_index, list_of_normalized_headers) or
    (None, None, None).
    """
    max_col = sheet.max_column
    if max_col < 13:
        return None, None, None

    for col_idx in range(1, max_col + 1):
        first_cell_val = normalize_header(sheet.cell(row=1, column=col_idx).value)
        second_cell_val = normalize_header(sheet.cell(row=2, column=col_idx).value)
        # Scan every non-empty cell matching the first expected header
        if first_cell_val == EXPECTED_17_HEADERS[0]:
            # Read next consecutive normalized values up to 17 columns
            row_headers = []
            for c in range(col_idx, min(col_idx + 17, max_col + 1)):
                row_headers.append(normalize_header(sheet.cell(row=1, column=c).value))

            # Phase 5: Check matching rules (All 17 match OR first 13 match)
            if len(row_headers) >= 17 and row_headers[:17] == EXPECTED_17_HEADERS:
                return 1, col_idx, row_headers[:17]
            elif len(row_headers) >= 13 and row_headers[:13] == EXPECTED_13_HEADERS:
                return 1, col_idx, row_headers[:13]
        elif second_cell_val == EXPECTED_17_HEADERS[0]:
            # Read next consecutive normalized values up to 17 columns from Row 2
            row_headers = []
            for c in range(col_idx, min(col_idx + 17, max_col + 1)):
                row_headers.append(normalize_header(sheet.cell(row=2, column=c).value))

            if len(row_headers) >= 17 and row_headers[:17] == EXPECTED_17_HEADERS:
                return 2, col_idx, row_headers[:17]
            elif len(row_headers) >= 13 and row_headers[:13] == EXPECTED_13_HEADERS:
                return 2, col_idx, row_headers[:13]

    return None, None, None


def extract_sheet_rows(sheet, header_row: int, start_col: int, header_length: int) -> list[list]:
    """
    Extracts row data starting after the header row down to sheet.max_row for header_length columns.
    Blank rows are ignored.
    """
    extracted_rows = []
    max_row = sheet.max_row

    for r in range(header_row + 1, max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(start_col, start_col + header_length)]
        # Check if any cell in this row slice contains non-empty value
        if any(v is not None and str(v).strip() != "" for v in row_vals):
            # Pad row if fewer than 17 items (e.g. 13-column match)
            if len(row_vals) < 17:
                row_vals.extend([None] * (17 - len(row_vals)))
            extracted_rows.append(row_vals)

    return extracted_rows


def process_data(file_paths: list[str], update_progress_callback=None) -> str:
    """
    Main entry point called by the Universal Excel Batch Processing GUI thread.
    """
    # Phase 1: Input Validation
    valid_files = validate_input_files(file_paths)

    # Pre-scan: Identify all visible worksheets across valid workbooks
    workbooks_to_process = []
    total_visible_sheets = 0
    warnings = []

    if update_progress_callback:
        update_progress_callback(0.0, "Scanning workbooks and visible worksheets...")

    for path in valid_files:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            visible_sheets = [sheet_name for sheet_name in wb.sheetnames if wb[sheet_name].sheet_state == 'visible']
            if visible_sheets:
                workbooks_to_process.append((path, visible_sheets))
                total_visible_sheets += len(visible_sheets)
            wb.close()
        except Exception as e:
            warnings.append(f"Could not open workbook {os.path.basename(path)}: {str(e)}")

    if total_visible_sheets == 0:
        raise ValueError("No visible worksheets found across the selected Excel files.")

    merged_data = []
    original_header_row = None
    original_header_cells = None
    processed_sheets_count = 0

    # Phase 2 & 7: Process Worksheets and Extract Data
    for file_path, sheet_names in workbooks_to_process:
        file_name = os.path.basename(file_path)
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in sheet_names:
                processed_sheets_count += 1
                progress = processed_sheets_count / total_visible_sheets

                if update_progress_callback:
                    update_progress_callback(
                        progress,
                        f"Reading: {file_name} | Sheet: {sheet_name}"
                    )

                sheet = wb[sheet_name]
                header_row, start_col, matched_headers = find_header_position(sheet)

                if header_row is None:
                    # Phase 8: Record warning for skipped worksheets
                    warnings.append(f"Skipped worksheet '{sheet_name}' in '{file_name}': Header not recognized.")
                    continue

                # Phase 6: Preserve Original Header from first valid sheet
                if original_header_row is None:
                    header_len = len(matched_headers)
                    original_header_row = [
                        sheet.cell(row=header_row, column=c).value for c in range(start_col, start_col + header_len)
                    ]
                    original_header_cells = [
                        sheet.cell(row=header_row, column=c) for c in range(start_col, start_col + header_len)
                    ]
                    # Pad header row to 17 items if 13-column format using default examples
                    if len(original_header_row) < 17:
                        original_header_row.extend(RAW_EXAMPLE_HEADERS[13:])
                        original_header_cells.extend([None] * (17 - len(original_header_cells)))

                # Extract Data
                rows = extract_sheet_rows(sheet, header_row, start_col, len(matched_headers))
                merged_data.extend(rows)

            wb.close()
        except Exception as e:
            warnings.append(f"Error processing workbook '{file_name}': {str(e)}")

    if original_header_row is None or not merged_data:
        raise ValueError("No worksheets matching the required PO header criteria were found.")

    # Phase 9 & 10: Create Output Workbook & Apply Cell/Date Formatting
    if update_progress_callback:
        update_progress_callback(0.95, "Generating merged result workbook...")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Merged Data"

    # Write Original Header and preserve the original header cell formatting
    for col_idx, header_value in enumerate(original_header_row, start=1):
        target_cell = out_ws.cell(row=1, column=col_idx, value=header_value)
        source_cell = original_header_cells[col_idx - 1] if original_header_cells and col_idx <= len(original_header_cells) else None
        copy_cell_style(source_cell, target_cell)

    # Determine date column indices from normalized standard headers (1-indexed for openpyxl)
    date_col_indices = [
        idx + 1 for idx, h_name in enumerate(EXPECTED_17_HEADERS) if h_name in DATE_HEADER_NAMES
    ]

    # Write Data Rows & Apply Formatting
    for row_idx, row_vals in enumerate(merged_data, start=2):
        out_ws.append(row_vals)
        for col_idx in range(1, 18):
            cell = out_ws.cell(row=row_idx, column=col_idx)
            
            # Apply date formatting to specified date columns
            if col_idx in date_col_indices:
                cell.number_format = 'dd/mm/yyyy'

    # Save Output File with Timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"Merged_Result_{timestamp}.xlsx"

    try:
        out_wb.save(output_filename)
    except Exception as e:
        raise IOError(f"Failed to save output Excel file: {str(e)}")

    # Print Warnings to console if any occurred
    if warnings:
        print("\n--- Processing Warnings ---")
        for warn in warnings:
            print(f"- {warn}")

    if update_progress_callback:
        update_progress_callback(1.0, "Exporting merged workbook complete!")

    return os.path.abspath(output_filename)